"""
XLSX OCR backend — openpyxl based extraction.

Each sheet becomes a Markdown section with a Markdown table.
Optional dependency: openpyxl.
"""

import logging
from typing import Any, Dict, Optional

from .base import OCRBackend

logger = logging.getLogger(__name__)

try:
    import openpyxl as _openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


class XlsxBackend(OCRBackend):
    """Extract content from XLSX/XLS files using openpyxl."""

    NAME = "xlsx"

    def is_available(self) -> bool:
        return _OPENPYXL_AVAILABLE

    def extract(self, file_path: str, options: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        if not _OPENPYXL_AVAILABLE:
            return self._result(file_path, error="XLSX backend requires 'openpyxl'")
        try:
            wb = _openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sections = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                sections.append(f"## {sheet_name}\n\n{self._rows_to_md(rows)}")
            wb.close()
            markdown = "\n\n".join(sections)
            return self._result(file_path, markdown=markdown, pages=len(wb.sheetnames))
        except Exception as exc:
            logger.warning("XlsxBackend failed for %s: %s", file_path, exc)
            return self._result(file_path, error=str(exc))

    def _rows_to_md(self, rows: list) -> str:
        def _cell(v) -> str:
            return str(v).replace("|", "\\|") if v is not None else ""

        lines = []
        for i, row in enumerate(rows):
            cells = [_cell(c) for c in row]
            lines.append("| " + " | ".join(cells) + " |")
            if i == 0:
                lines.append("| " + " | ".join(["---"] * len(cells)) + " |")
        return "\n".join(lines)
